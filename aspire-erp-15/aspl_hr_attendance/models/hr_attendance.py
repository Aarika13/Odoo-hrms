# -*- coding: utf-8 -*-

import time
from odoo.exceptions import UserError, ValidationError
from odoo.tools import DEFAULT_SERVER_DATE_FORMAT, DEFAULT_SERVER_DATETIME_FORMAT
from odoo import tools, models, fields, api, _
import openpyxl
from datetime import datetime, time, timedelta, date
import traceback
import pytz
import logging
from collections import OrderedDict
import copy
import base64
import io
from dateutil.relativedelta import relativedelta

_logger = logging.getLogger(__name__)


class IrRule(models.Model):
    _inherit = 'ir.rule'

    @api.model
    def archive_attendance_rule(self):
        if self.env.ref('hr_attendance.hr_attendance_rule_attendance_manager', raise_if_not_found=False):
            self.env.ref('hr_attendance.hr_attendance_rule_attendance_manager').update({'active': False})


class HrAttendance(models.Model):
    _inherit = ['hr.attendance']
    comment = fields.Text('Details')
    file_id = fields.Many2one('attendance.biometric.file', 'File', readonly=True)
    atten_status = fields.Boolean("Status")
    has_error = fields.Boolean("Mistake")

    def emp_attendance_time_scheduler(self):
        """ Employee attendance scheduler """
        attend_file_obj = self.env['attendance.biometric.file'].search([])
        if self.ids:
            attendanceSheetDataObj = attend_file_obj.browse(self.ids)
        else:
            att_file_obj = self.env['attendance.biometric.file'].search([('id', '=', self.file_id.id)])
            attenExcSheetDataObjSearch = att_file_obj.search([])
            attendanceSheetDataObj = att_file_obj.browse(attenExcSheetDataObjSearch)

        count = 0
        try:
            attend_file_obj.write({'status': 'in_progress'})
            # iterate attendance biometric file record
            for record in attend_file_obj:
                file_id = record.id
                time = datetime.now()
                # check file status is failure
                if 'failure' in str(record.status) or str(record.status).upper() in 'FAILURE':
                    count += 1
                    self.delete_hr_attendance(file_id)
                    # self.delete_attendance_daily_summary(file_id)
                    # self.delete_attendance_monthly_summary(file_id)
                    self.update_attendance_file_status_is_in_progress(time, file_id)
                    self.excel_sheet_emp_attendance_record(time, file_id)
            # call first time if file record is not created
            if count == 0:
                self.update_attendance_file_status_is_in_progress(time, file_id)
                self.excel_sheet_emp_attendance_record(time, file_id)
        except Exception as e:
            _logger.error('Something is wrong')
            _logger.error(str(e))

    def delete_hr_attendance(self, file_id):
        """ delete hr attendance record """
        try:
            attend_rec_search_obj = self.search([('file_id', '=', file_id)])
            attend_rec_search_obj.unlink()
            return True
        except Exception as e:
            _logger.error('Something is wrong')
            _logger.error(str(e))
            return False

    # def delete_attendance_daily_summary(self, file_id):
    #     """ delete attendance daily summary record """
    #     attend_record_obj = self.env['attendance.daily.summary'].search([('file_id', '=', file_id)])
    #     try:
    #         attend_record_obj.unlink()
    #         return True
    #     except Exception as e:
    #         _logger.error('Something is wrong')
    #         _logger.error(str(e))
    #         return False

    # def delete_attendance_monthly_summary(self, file_id):
    #     """ delete attendance monthly summary record """
    #
    #     attend_record_obj = self.env['attendance.monthly.summary'].search([('file_id', '=', file_id)])
    #     try:
    #         attend_record_obj.unlink()
    #         return True
    #     except Exception as e:
    #         _logger.error('Something is wrong')
    #         _logger.error(str(e))
    #         return False

    def _employee_get(self):
        """ get employee id """
        hr_employee_obj = self.env['hr.employee'].search([('user_id', '=', self.env.user.id)])
        if hr_employee_obj:
            return hr_employee_obj
        return False

    def utc_to_local(self, date_time, file_id):
        """ convert utc to local timezone """
        try:
            tz = pytz.timezone(self.env.user.partner_id.tz)
            local_dt = date_time.replace(tzinfo=pytz.utc).astimezone(tz)
            return tz.normalize(local_dt)
        except:
            emp_id = self._employee_get()
            if emp_id.work_email:

                message = (
                        "Hi " + emp_id.name
                        + ", ""\n\n There was an issue while proces Sign attendance records. Below is the detail in brief:"
                          "\n\n Timezone is not set. Make sure "
                        + emp_id.name + " has timezone assigned.")
                vals = {
                    'subject': 'Issue in Attendance Record',
                    'body_html': '<pre>%s</pre>' % message,
                    'email_to': emp_id.work_email,
                    'email_from': 'pgandhi@aspiresoftware.in',
                }
                email_ids = self.env['mail.mail'].create(vals)
                if email_ids:
                    email_ids.send()
            end_time = datetime.now()
            self.update_attendance_file_status_is_failure(file_id, end_time)

    def split(self, str, chunk_size):
        return [str[i:i + chunk_size] for i in range(0, len(str), chunk_size)]

    def split_excel_data(self, empid, date_obj, data, file_id):
        """ split attendance excel file data """
        dict_obj = OrderedDict()
        try:
            split_excel_data = str(data).split('(')
            time = split_excel_data[0].split(':')
            split_type = split_excel_data[1].split(')')
            func_type = split_type[0]
            date_time = datetime.strptime(str(date_obj), DEFAULT_SERVER_DATE_FORMAT) + timedelta(
                hours=int(time[0])) + timedelta(minutes=int(time[1]))
            data_date = self.utc_to_local(date_time, file_id)
            list = data_date.strftime('%z')
            if '-' in list:
                str_list = list.lstrip('-')
                data_list = self.split(str_list, 2)
            else:
                str_list = list.lstrip('+')
                data_list = self.split(str_list, 2)

            if '-' in list:
                date_time = date_time - (-timedelta(hours=int(data_list[0]))) - (-timedelta(minutes=int(data_list[1])))
            else:
                date_time = date_time - timedelta(hours=int(data_list[0])) - timedelta(minutes=int(data_list[1]))

            hr_employee_obj = self.env['hr.employee'].search([])
            count = 0
            emp_num = True if empid.isnumeric() else False
            for record in hr_employee_obj:
                if record.biometric_no and (
                        (emp_num and record.biometric_no.isnumeric() and int(record.biometric_no) == int(empid)) or (
                        empid == record.biometric_no)):
                    employee_id = record.id
                    emp_name = record.name
                    count += 1
            if count == 1:
                dict_obj['empid'] = employee_id
                dict_obj['dateTime'] = date_time
                dict_obj['type'] = str(func_type)
                dict_obj['time'] = str(time[0]) + '.' + str(time[1])
                return dict_obj

        except Exception as e:
            _logger.error('Something is wrong')
            _logger.error(str(e))
            end_time = datetime.now()
            self.update_attendance_file_status_is_failure(file_id, end_time)
            return False

    def update_attendance_file_status_is_failure(self, file_id, end_time):
        """ update attendance file status is failure """

        att_file_obj = self.env['attendance.biometric.file'].search([('id', '=', file_id)])
        try:
            att_file_obj.write({
                'end_time': end_time,
                'status': 'failure'})
        except Exception as e:
            _logger.error('Something is wrong')
            _logger.error(str(e))
            return False

    def update_attendance_file_status_is_in_progress(self, start_time, file_id):
        """ update attendance file status is In progress """
        att_file_obj = self.env['attendance.biometric.file'].search([('id', '=', file_id)])
        try:
            att_file_obj.write({'start_time': start_time,
                                'status': 'in_progress'})
        except Exception as e:
            _logger.error('Something is wrong')
            _logger.error(str(e))
            return False

    def excel_sheet_emp_attendance_record(self, time, file_id):
        """ process attendance excel sheet record """
        try:
            att_file_obj = self.env['attendance.biometric.file'].search([('id', '=', file_id)])
            for i in att_file_obj:
                dataa = base64.b64decode(i.datas)
                xls_file_like = io.BytesIO(dataa)
                workbook = openpyxl.load_workbook(xls_file_like)
                sheet = workbook[workbook.sheetnames[0]]
                start_row = 2
                start_time = datetime.now()
                timesheet_data_list = []
                # update attendance file status is In progress
                self.update_attendance_file_status_is_in_progress(start_time, file_id)
                # process attendance excel sheet one by one row
                for i in range(start_row, sheet.max_row + 1):
                    start_column = 5
                    temp_emp_id = sheet.cell(row=i, column=2).value
                    if temp_emp_id == ' ' or str(temp_emp_id) not in '..':
                        emp_id = sheet.cell(row=i, column=2).value
                    date_obj = sheet.cell(row=i, column=4).value
                    hr_employee_obj = self.env['hr.employee'].search([])
                    count = 0
                    emp_num = True if emp_id.isnumeric() else False
                    for record in hr_employee_obj:
                        if record.biometric_no and ((emp_num and record.biometric_no.isnumeric() and int(
                                record.biometric_no) == int(emp_id)) or (record.biometric_no == emp_id)):
                            employee_id = record.id
                            emp_name = record.name
                            count += 1
                    if count == 1:
                        pass
                    else:
                        continue
                    time_sheet_dict = {}
                    time_sheet_dict['file_id'] = file_id
                    # process attendance excel sheet row , one by one column
                    while start_column <= sheet.max_column:
                        data = sheet.cell(row=i, column=start_column).value
                        if data:
                            data_list = self.split_excel_data(emp_id, date_obj, data, id)
                            if data_list:
                                timesheet_data_list.append(data_list)
                            start_column += 1
                        else:
                            break
                    count = 0
                hr_employee_obj = self.env['hr.employee'].search([])
                for record in hr_employee_obj:
                    emp_attend_list = []
                    for items in timesheet_data_list:
                        if 'empid' in items:
                            if items['empid'] == record.id:
                                emp_attend_list.append(items)
                    if emp_attend_list:
                        self.remove_duplicate_attendance_swipe_record(emp_attend_list, time_sheet_dict, date_obj,
                                                                      file_id)
                end_time = datetime.now()
                att_file_obj.write({
                    'end_time': end_time,
                    'status': 'success'})
                # self.cr.execute(
                #     "update hr_attendance set sheet_id=hr_timesheet_sheet_sheet.id from hr_timesheet_sheet_sheet where hr_attendance.employee_id = hr_timesheet_sheet_sheet.employee_id and date_trunc('day',hr_attendance.name) between hr_timesheet_sheet_sheet.date_from and hr_timesheet_sheet_sheet.date_to and hr_attendance.sheet_id is null")
        # self.pool.get('attendance.daily_summary').attendance_daily_summary_scheduler(cr,uid,None,None,file_id,context=None)
        # self.pool.get('attendance.monthly.summary').attendance_monthly_summary_scheduler(cr,uid,file_ids=None,context=None)
        except Exception as e:

            _logger.error('Something is wrong')
            _logger.error(str(e))
            self.cr.rollback()
            end_time = datetime.now()
            # update hr attendance record status is failure
            self.update_attendance_file_status_is_failure(file_id, end_time)
            raise ValidationError('Incorrect Excel Time Sheet!. Please check and correct Excel Sheet.')

    def generate_attendance_sequence(self, emp_id, wfh_date):
        list_in_time = []
        list_out_time = []
        d = wfh_date + " " + "00:00:00"
        single_date = datetime.strptime(d, DEFAULT_SERVER_DATETIME_FORMAT)
        tomorrow = single_date + timedelta(days=1)
        tomorrow_beginning = datetime(tomorrow.year, tomorrow.month, tomorrow.day, 0, 0, 0, 0)
        tmb = tomorrow_beginning.strftime("%Y-%m-%d %H:%M:%S")
        today = single_date
        today_beginning = datetime(today.year, today.month, today.day, 0, 0, 0, 0)
        tb = today_beginning.strftime("%Y-%m-%d %H:%M:%S")

        hr_attendance_obj = self.search(
            [('employee_id', '=', emp_id), ('name', '>=', tb), ('name', '<', tmb)],
            order="name asc")
        for ii in hr_attendance_obj:
            if str(ii['type']) in 'I':
                list_in_time.append(ii)
            else:
                list_out_time.append(ii)
        final_list = []
        in_count = 0
        out_count = 0
        count = 0
        for ii in range(0, len(hr_attendance_obj)):
            if in_count == 0 and count == 0 and len(list_in_time) > 0:
                final_list.append(list_in_time[ii])
                in_count += 1
                count += 1

            elif count != 0 and final_list[-1:][0]['type'] == 'I':
                if out_count < len(list_out_time):
                    if final_list[-1:][0]['dateTime'] <= list_out_time[out_count]['dateTime']:
                        final_list.append(list_out_time[out_count])
                        count += 1
                    else:
                        if final_list[-2:-1] and list_out_time[out_count]['time'] != final_list[-2:-1][0]['time']:
                            final_list[-2:-1][0]['time'] = str(final_list[-2:-1][0]['time']) + " , " + str(
                                list_out_time[out_count]['time'])
                        pass
                    out_count += 1

            elif count != 0 and final_list[-1:][0]['type'] == 'O':
                flag = False
                if in_count < len(list_in_time):
                    date_time = list_in_time[in_count]['dateTime']
                    list_date_time = self.utc_to_local(date_time, file_id=1)
                    date_time = final_list[-2:][0]['dateTime']
                    final_date_time = self.utc_to_local(date_time, file_id=1)
                    if final_list[-1:][0]['dateTime'] <= list_in_time[in_count]['dateTime']:
                        final_list.append(list_in_time[in_count])
                        in_count += 1
                        count += 1
                    elif final_list[-1:][0]['dateTime'] > list_in_time[in_count]['dateTime']:
                        if final_list[-2:-1][0]['time'] != list_in_time[in_count]['time']:
                            if list_date_time.date() == final_date_time.date():
                                list_in_time[in_count]['time'] = str(final_list[-2:-1][0]['time']) + " , " + str(
                                    list_in_time[in_count]['time'])
                            else:
                                flag = True
                        del final_list[-2:-1]
                        final_list.insert(-1, list_in_time[in_count])
                        if flag:
                            final_list.insert(-2, list_in_time[in_count - 1])
                            copy_list = copy.deepcopy(list_in_time)
                            copy_list[in_count - 1]['dateTime'] += timedelta(seconds=1)
                            copy_list[in_count - 1]['type'] = 'O'
                            final_list.insert(-2, copy_list[in_count - 1])
                        in_count += 1
                    else:
                        pass
                        in_count += 1
        data_list = []
        count = 0
        tz = pytz.timezone(self.env.user.partner_id.tz)
        # add timedelta in all records with same datetime
        for r in final_list:
            if r['type'] == 'I' and count == 0:
                data_list.append(r)
                count += 1
            if r['type'] == 'O' and count == 0:
                data_list.append(r)
                count += 1
            elif count != 0 and data_list[-1:][0]['type'] == 'I' and r['type'] == 'O':
                if tz.normalize(
                        data_list[-1:][0]['dateTime'].replace(tzinfo=pytz.utc).astimezone(tz)).date() == tz.normalize(
                    r['dateTime'].replace(tzinfo=pytz.utc).astimezone(tz)).date():
                    pass
                if data_list[-1:][0]['dateTime'] == r['dateTime']:
                    r['dateTime'] += timedelta(seconds=1)
                    data_list.append(r)
                    for ii in range(count, len(final_list)):
                        if final_list[ii]['dateTime'] == data_list[-2:-1][0]['dateTime']:
                            final_list[ii]['dateTime'] += timedelta(seconds=1)
                if tz.normalize(
                        data_list[-1:][0]['dateTime'].replace(tzinfo=pytz.utc).astimezone(tz)).date() != tz.normalize(
                    r['dateTime'].replace(tzinfo=pytz.utc).astimezone(tz)).date() and tz.normalize(
                    r['dateTime'].replace(tzinfo=pytz.utc).astimezone(tz)) > tz.normalize(
                    r['dateTime'].replace(tzinfo=pytz.utc).astimezone(tz)).replace(hour=6, minute=0, second=0,
                                                                                   microsecond=0):
                    tdelta = time(hour=2, minute=30, second=0)
                    if r['dateTime'] > datetime.combine(datetime.strptime(str(r['dateTime'].date()), '%Y-%m-%d').date(),
                                                        tdelta):
                        r['dateTime'] = data_list[-1:][0]['dateTime'] + timedelta(seconds=1)
                    data_list.append(r)
                else:
                    data_list.append(r)
                count += 1
            elif count != 0 and data_list[-1:][0]['type'] == 'O' and r['type'] == 'I':
                if tz.normalize(
                        data_list[-1:][0]['dateTime'].replace(tzinfo=pytz.utc).astimezone(tz)).date() == tz.normalize(
                    r['dateTime'].replace(tzinfo=pytz.utc).astimezone(tz)).date():
                    pass
                if data_list[-1:][0]['dateTime'] == r['dateTime']:
                    r['dateTime'] += timedelta(seconds=1)
                    data_list.append(r)
                    for ii in range(count, len(final_list)):
                        if final_list[ii]['dateTime'] == data_list[-2:-1][0]['dateTime']:
                            final_list[ii]['dateTime'] += timedelta(seconds=1)
                else:
                    data_list.append(r)
                count += 1

        # if last sign-out not found
        if count != 0 and data_list[-1:][0]['type'] == 'I':
            data_list.append(
                dict(empid=data_list[-1:][0]['empid'], dateTime=data_list[-1:][0]['dateTime'] + timedelta(seconds=1),
                     type='O', time=data_list[-1:][0]['time']))
        attendance_list = []
        for items in data_list:
            if items not in attendance_list:
                attendance_list.append(items)

        self.attendance_record(attendance_list, time_sheet_dict=None)

    def remove_duplicate_attendance_swipe_record(self, attendance_list, time_sheet_dict, dateObj, file_id):
        list_in_time = []
        list_out_time = []
        for ii in attendance_list:
            if str(ii['type']) in 'I':
                list_in_time.append(ii)
            else:
                list_out_time.append(ii)

        final_list = []
        in_count = 0
        out_count = 0
        count = 0
        for ii in range(0, len(attendance_list)):
            if in_count == 0 and count == 0 and len(list_in_time) > 0:
                final_list.append(list_in_time[ii])
                in_count += 1
                count += 1
            elif count != 0 and final_list[-1:][0]['type'] == 'I':
                if out_count < len(list_out_time):
                    if final_list[-1:][0]['dateTime'] <= list_out_time[out_count]['dateTime']:
                        final_list.append(list_out_time[out_count])
                        count += 1
                    else:
                        if final_list[-2:-1] and list_out_time[out_count]['time'] != final_list[-2:-1][0]['time']:
                            final_list[-2:-1][0]['time'] = str(final_list[-2:-1][0]['time']) + " , " + str(
                                list_out_time[out_count]['time'])
                        pass
                    out_count += 1
            elif count != 0 and final_list[-1:][0]['type'] == 'O':
                flag = False
                if in_count < len(list_in_time):
                    date_time = list_in_time[in_count]['dateTime']
                    list_date_time = self.utc_to_local(date_time, file_id)
                    date_time = final_list[-2:][0]['dateTime']
                    final_date_time = self.utc_to_local(date_time, file_id)
                    if final_list[-1:][0]['dateTime'] <= list_in_time[in_count]['dateTime']:
                        final_list.append(list_in_time[in_count])
                        in_count += 1
                        count += 1
                    elif final_list[-1:][0]['dateTime'] > list_in_time[in_count]['dateTime']:
                        if final_list[-2:-1][0]['time'] != list_in_time[in_count]['time']:
                            if list_date_time.date() == final_date_time.date():
                                list_in_time[in_count]['time'] = str(final_list[-2:-1][0]['time']) + " , " + str(
                                    list_in_time[in_count]['time'])
                            else:
                                flag = True
                        del final_list[-2:-1]
                        final_list.insert(-1, list_in_time[in_count])
                        if flag:
                            final_list.insert(-2, list_in_time[in_count - 1])
                            copy_list = copy.deepcopy(list_in_time)
                            copy_list[in_count - 1]['dateTime'] += timedelta(seconds=1)
                            copy_list[in_count - 1]['type'] = 'O'
                            final_list.insert(-2, copy_list[in_count - 1])
                        in_count += 1
                    else:
                        pass
                        in_count += 1
        data_list = []
        count = 0
        tz = pytz.timezone(self.env.user.partner_id.tz)
        # add timedelta in all records with same dateTime
        for r in final_list:
            if r['type'] == 'I' and count == 0:
                data_list.append(r)
                count += 1
            if r['type'] == 'O' and count == 0:
                data_list.append(r)
                count += 1
            elif count != 0 and data_list[-1:][0]['type'] == 'I' and r['type'] == 'O':
                if tz.normalize(
                        data_list[-1:][0]['dateTime'].replace(tzinfo=pytz.utc).astimezone(tz)).date() == tz.normalize(
                    r['dateTime'].replace(tzinfo=pytz.utc).astimezone(tz)).date():
                    pass
                if data_list[-1:][0]['dateTime'] == r['dateTime']:
                    r['dateTime'] += timedelta(seconds=1)
                    data_list.append(r)
                    for ii in range(count, len(final_list)):
                        if final_list[ii]['dateTime'] == data_list[-2:-1][0]['dateTime']:
                            final_list[ii]['dateTime'] += timedelta(seconds=1)
                if tz.normalize(
                        data_list[-1:][0]['dateTime'].replace(tzinfo=pytz.utc).astimezone(tz)).date() != tz.normalize(
                    r['dateTime'].replace(tzinfo=pytz.utc).astimezone(tz)).date() and tz.normalize(
                    r['dateTime'].replace(tzinfo=pytz.utc).astimezone(tz)) > tz.normalize(
                    r['dateTime'].replace(tzinfo=pytz.utc).astimezone(tz)).replace(hour=6, minute=0, second=0,
                                                                                   microsecond=0):
                    tdelta = time(hour=2, minute=30, second=0)
                    if r['dateTime'] > datetime.combine(datetime.strptime(str(r['dateTime'].date()), '%Y-%m-%d').date(),
                                                        tdelta):
                        r['dateTime'] = data_list[-1:][0]['dateTime'] + timedelta(seconds=1)
                    data_list.append(r)
                else:
                    data_list.append(r)
                count += 1
            elif count != 0 and data_list[-1:][0]['type'] == 'O' and r['type'] == 'I':
                if tz.normalize(
                        data_list[-1:][0]['dateTime'].replace(tzinfo=pytz.utc).astimezone(tz)).date() == tz.normalize(
                    r['dateTime'].replace(tzinfo=pytz.utc).astimezone(tz)).date():
                    # r['duration'] = r['dateTime'] - dataList[-1:][0]['dateTime']
                    # (h, m, s) = str(r['duration']).split(':')
                    # r['duration_decimal'] = int(h) * 3600 + int(m) * 60 + int(s)
                    pass
                if data_list[-1:][0]['dateTime'] == r['dateTime']:
                    r['dateTime'] += timedelta(seconds=1)
                    data_list.append(r)
                    for ii in range(count, len(final_list)):
                        if final_list[ii]['dateTime'] == data_list[-2:-1][0]['dateTime']:
                            final_list[ii]['dateTime'] += timedelta(seconds=1)
                else:
                    data_list.append(r)
                count += 1

        # if last sign-out not found
        if count != 0 and data_list[-1:][0]['type'] == 'I':
            data_list.append(
                dict(empid=data_list[-1:][0]['empid'], dateTime=data_list[-1:][0]['dateTime'] + timedelta(seconds=1),
                     type='O', time=data_list[-1:][0]['time']))
        attendance_list = []
        for items in data_list:
            if items not in attendance_list:
                attendance_list.append(items)
        self.attendance_record(attendance_list, time_sheet_dict)

    def attendance_record(self, attendance_list, time_sheet_dict):
        """ process excel file attendance record """
        file_id = ' '
        try:
            att_final_list = []
            pos = 0
            for record in attendance_list:
                temp_type = record['type']
                file_id = time_sheet_dict['file_id']
                temp_emp_id = record['empid']
                temp_date_time = str(record['dateTime'])
                comment = str(record['time'])
                duration = None
                duration_decimal = None
                if 'duration' in record:
                    duration = record['duration']
                if 'duration_decimal' in record:
                    duration_decimal = (record['duration_decimal'] / 60)
                if temp_type == 'I':
                    att_final_list.append({
                        'emp_id': temp_emp_id,
                        'check_in': temp_type,
                        'in_time': temp_date_time,
                        'comment': comment
                    })
                    # if len(comment) > 5: as per old file
                    if len(comment) > 6:
                        att_final_list[pos]['error'] = True
                    else:
                        att_final_list[pos]['error'] = False
                else:
                    att_final_list[pos]['check_out'] = temp_type
                    att_final_list[pos]['out_time'] = temp_date_time
                    pos += 1
            pos = 0
            flag = True
            err_list_string = ""
            print("att_final_list", att_final_list)
            for record in att_final_list:
                print("record", record)
                # Attendance Create
                try:
                    attendance = self.env['hr.attendance'].create({
                        'check_in': record['in_time'],
                        'check_out': record['out_time'],
                        'employee_id': record['emp_id'],
                        'comment': record['comment'],
                        'file_id': time_sheet_dict['file_id'],
                    })
                except Exception as e:
                    flag = False
                    print("record['in_time']", record['in_time'])
                    print("record['out_time']", record['out_time'])
                    err_list_string = err_list_string + "\n" + str(e)
                    _logger.error(traceback.format_exc())
                    _logger.error(err_list_string)
                    # end_time = datetime.now()
                    # self.update_attendance_file_status_is_failure(file_id, end_time)
                    continue

                if record['error']:
                    attendance.write({'has_error': record['error']})
                    print(" record['comment']", record['comment'])
                    print("record['in_time']", attendance.check_in)
                    print("record['out_time']", attendance.check_out)
                    print("record['error']", record['error'])

            search_obj = self.env['attendance.biometric.file'].browse(file_id)
            _logger.info("Searched file id is: [%s][%s]", search_obj, type(search_obj))
            if search_obj and file_id != ' ':
                if flag == True:
                    display_msg = "File processing completed. No issues to report."
                    search_obj.message_post(body=display_msg)
                elif flag == False:
                    display_msg = "File processing completed. Following issues were found during processing which need to be handled manually:" + "\n" + err_list_string
                    search_obj.message_post(body=display_msg)

        except Exception as e:
            _logger.error(traceback.format_exc())
            _logger.error(str(e))
            end_time = datetime.now()
            # update hr attendance record status is failure
            self.update_attendance_file_status_is_failure(file_id, end_time)
            raise ValidationError('Incorrect Excel Time Sheet!. Please check and correct Excel Sheet.')

    def add_auto_attendance_from_biometric(self):
        day_before_current_date = date.today() - relativedelta(days=1)
        # date_start = day_before_current_date.replace(hour=0, minute=0, second=0, microsecond=0)
        attendance_record = self.env['attendance.wizard'].create({
            'start_date': day_before_current_date,
            'end_date': day_before_current_date
        })
        connection_data = self.env['connector.sqlserver'].search([('state', '=', 'active')])
        for connection in connection_data:
            if connection.auto_gen_attendance:
                context = {
                    'connection_id': connection.id
                }
                attendance_record.with_context(context).generate_attendance()

    # def create_attendance_record(self, file_id, duration, duration_decimal, temp_type, temp_date_time,
    #                              temp_emp_id, comment):
    #     """ create new attendance record """
    #
    #     action = ''
    #     try:
    #         # Deprecated below object in Odoo 15
    #         # if temp_type in 'I':
    #         #     search_domain = [('name', '=', 'Sign in'), ]
    #         #     stage_ids = self.pool.get('hr.action.reason').search(cr, uid, search_domain, context=context)
    #         #     action = 'sign_in'
    #         #     action_desc = stage_ids[0]
    #         # else:
    #         #     search_domain = [('name', '=', 'Sign out'), ]
    #         #     stage_ids = self.pool.get('hr.action.reason').search(cr, uid, search_domain, context=context)
    #         #     action = 'sign_out'
    #         #     action_desc = stage_ids[0]
    #
    #         # sheet_id = None
    #         # sheet_obj = self.env('hr_timesheet_sheet.sheet').search([])
    #         employee_data = self.env['hr.employee'].search([('id', '=', temp_emp_id)])
    #
    #         # if sheet_obj:
    #         #     for record in sheet_obj:
    #         #         if datetime.strptime(record.date_from,
    #         #                              DEFAULT_SERVER_DATE_FORMAT).date() <= temp_date_time.date() and datetime.strptime(
    #         #             record.date_to,
    #         #             DEFAULT_SERVER_DATE_FORMAT).date() >= temp_date_time.date() and record.user_id.id == employee_data.user_id.id:
    #         #             sheet_id = record.id
    #         #         else:
    #         #             sheet_id = None
    #         error = False
    #         if len(comment.lstrip()) > 6:
    #             error = True
    #         date = temp_date_time.strftime('%Y-%m-%d %H:%M:%S')
    #
    #     except ValidationError as v:
    #
    #         if str(v) in "(u'Error ! Sign in (resp. Sign out) must follow Sign out (resp. Sign in)', None)":
    #             pass
    #         else:
    #
    #             _logger.error(traceback.format_exc())
    #             _logger.error(str(v))
    #             end_time = datetime.now()
    #             # update hr attendance record status is failure
    #             self.update_attendance_file_status_is_failure(file_id, end_time)
    #
    #
    #     except Exception as e:
    #         _logger.error('Something is wrong')
    #         _logger.error(traceback.format_exc())
    #         _logger.error(str(e))
    #
    #         end_time = datetime.now()
    #         # update hr attendance record status is failure
    #         self.update_attendance_file_status_is_failure(file_id, end_time)

    # Commented by RPJ : 29-12-2021
    # def write(self, vals):
    #     res = super(HrAttendance, self).write(vals)
    #     name = self.name
    #     emp_id = self.employee_id.id
    #     # self.update_daily_summary(cr, uid, ids,name,emp_id,context=None)
    #     # self.update_monthly_summary(cr, uid, ids,name,emp_id, context=None)
    #     return res
    # Commented by RPJ : 29-12-2021
    # def create(self, data):
    #     res = super(HrAttendance, self).create(data)
    #     if 'name' in data and 'employee_id' in data and data['action'] == 'sign_out' and 'file_id' not in data:
    #         name = data['name']
    #         emp_id = data['employee_id']
    #         start_date = datetime.strptime(name, DEFAULT_SERVER_DATETIME_FORMAT).strftime('%Y-%m-%d')
    #         self.env.ref('attendance.daily.summary').attendance_daily_summary_scheduler(emp_id, start_date,
    #                                                                                     file_id=None)
    #     # self.update_daily_summary(cr, uid, ids,name,emp_id ,context=None)
    #     # self.update_monthly_summary(cr, uid, ids,name,emp_id ,context=None)
    #     if 'file_id' in data and 'name' in data and 'employee_id' in data:
    #         name = data['name']
    #         emp_id = data['employee_id']
    #         update_sql = 'UPDATE public.attendance_daily_summary SET "dailySumm_status"= False WHERE EXTRACT(MONTH FROM date)=' + str(
    #             datetime.strptime(name,
    #                               DEFAULT_SERVER_DATETIME_FORMAT).month) + ' and EXTRACT(YEAR FROM date) = ' + str(
    #             datetime.strptime(name, DEFAULT_SERVER_DATETIME_FORMAT).year) + ' and emp_id = ' + str(emp_id) + ';'
    #         self.cr.execute(update_sql)
    #         delete_sql = 'DELETE FROM public.attendance_monthly_summary WHERE EXTRACT(MONTH FROM month) = ' + str(
    #             datetime.strptime(name,
    #                               DEFAULT_SERVER_DATETIME_FORMAT).month) + ' and EXTRACT(YEAR FROM month) = ' + str(
    #             datetime.strptime(name, DEFAULT_SERVER_DATETIME_FORMAT).year) + ' and employee = ' + str(emp_id) + ';'
    #         self.cr.execute(delete_sql)
    #     return res

    # def unlink(self):
    #     for attendance in self:
    #         name = attendance.name
    #         emp_id = attendance.employee_id.id
    #     res = super(HrAttendance, self).unlink()
    #     # self.update_daily_summary(cr, uid, ids,name,emp_id,context=None)
    #     # self.update_monthly_summary(cr, uid, ids,name,emp_id, context=None)
    #     return res

    # Not used in the file: commented in all the above methods
    # def update_monthly_summary(self, name, emp_id):
    #
    #     updateSql = 'UPDATE public.attendance_daily_summary SET "dailySumm_status"= False WHERE EXTRACT(MONTH FROM date)=' + str(
    #         datetime.strptime(name, DEFAULT_SERVER_DATETIME_FORMAT).month) + ' and EXTRACT(YEAR FROM date) = ' + str(
    #         datetime.strptime(name, DEFAULT_SERVER_DATETIME_FORMAT).year) + ' and emp_id = ' + str(emp_id) + ';'
    #     self.cr.execute(updateSql)
    #
    #     deleteSql = 'DELETE FROM public.attendance_monthly_summary WHERE EXTRACT(MONTH FROM month) = ' + str(
    #         datetime.strptime(name, DEFAULT_SERVER_DATETIME_FORMAT).month) + ' and EXTRACT(YEAR FROM month) = ' + str(
    #         datetime.strptime(name, DEFAULT_SERVER_DATETIME_FORMAT).year) + ' and employee = ' + str(emp_id) + ';'
    #
    #     self.cr.execute(deleteSql)
    #
    #     self.env.ref('attendance.monthly.summary').attendance_monthly_summary_scheduler(file_ids=None)

    # Not used in the file: commented in all the above methods
    # def update_daily_summary(self, cr, uid, ids, name, emp_id, context=None):
    #
    #     name = datetime.strftime(datetime.strptime(name, DEFAULT_SERVER_DATETIME_FORMAT),
    #                              DEFAULT_SERVER_DATETIME_FORMAT)
    #     hrAttend = self.pool.get('hr.attendance')
    #     yesterday = datetime.strptime(name, DEFAULT_SERVER_DATETIME_FORMAT) + timedelta(days=1)
    #     yesterday_beginning = datetime(yesterday.year, yesterday.month, yesterday.day, 0, 0, 0, 0)
    #     yb = yesterday_beginning.strftime("%Y-%m-%d %H:%M:%S")
    #     today = datetime.strptime(name, DEFAULT_SERVER_DATETIME_FORMAT)
    #     today_beginning = datetime(today.year, today.month, today.day, 0, 0, 0, 0)
    #     tb = today_beginning.strftime("%Y-%m-%d %H:%M:%S")
    #
    #     hrAttendSerchDateObj = hrAttend.search(cr, uid,
    #                                            [('employee_id', '=', emp_id), ('name', '>=', tb), ('name', '<', yb)],
    #                                            order="name asc", context=context)
    #     hrAttendDataDateObj = hrAttend.browse(cr, uid, hrAttendSerchDateObj, context=context)
    #     signInTime = None
    #     for attenRecord in hrAttendDataDateObj:
    #         if attenRecord.action == 'sign_in':
    #             signInTime = attenRecord.name
    #         elif attenRecord.action == 'sign_out' and signInTime is not None:
    #             workedhours_datetime = (
    #                     datetime.strptime(attenRecord.name, '%Y-%m-%d %H:%M:%S') - datetime.strptime(signInTime,
    #                                                                                                  '%Y-%m-%d %H:%M:%S'))
    #             new_worked_hours = ((workedhours_datetime.seconds) / 60) / 60.0
    #             cr.execute(""" UPDATE public.hr_attendance SET worked_hours= %s WHERE id = %s;""" % (
    #                 new_worked_hours, attenRecord.id))
    #     if hrAttendDataDateObj:
    #         updateSql = 'UPDATE public.hr_attendance SET atten_status = False WHERE id in' + str(
    #             hrAttendSerchDateObj).replace("[", "(").replace("]", ")") + ';'
    #         cr.execute(updateSql)
    #
    #         attend_rec_sum_obj = self.env['attendance.daily.summary'].search(
    #             [('emp_id', '=', emp_id), ('date', '=', today_beginning.date())])
    #         attend_rec_sum_obj.unlink()
    #
    #         self.env.ref('attendance.daily.summary').attendance_daily_summary_scheduler(emp_id=None,
    #                                                                                     start_date=None, file_id=None)

    # Deprecated in Odoo 15
    # class hr_action_reason(osv.osv):
    #     _inherit = "hr.action.reason"
    #
    #     _columns = {
    #         'sequence': fields.integer('Sequence'),
    #
    #     }